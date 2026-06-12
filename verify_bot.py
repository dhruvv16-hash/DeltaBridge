import unittest
import math
import hmac
import hashlib
import time
from unittest.mock import MagicMock, patch
from delta_client import DeltaClient
from config import Config

class TestDeltaBot(unittest.TestCase):
    def setUp(self):
        self.api_key = "test_key"
        self.api_secret = "test_secret"
        self.client = DeltaClient(
            api_key=self.api_key,
            api_secret=self.api_secret,
            base_url="https://api.delta.exchange"
        )

    def test_symbol_normalization(self):
        """Verify symbols are cleaned up correctly."""
        # Test clean symbol extraction
        prod1 = {"symbol": "ETHUSD", "id": 27}
        prod2 = {"symbol": "BTCUSD", "id": 1}
        self.client._products_cache = [prod1, prod2]
        self.client._products_cache_time = time.time()
        
        self.assertEqual(self.client.get_product_by_symbol("ETHUSD.P")["id"], 27)
        self.assertEqual(self.client.get_product_by_symbol("ethusd.p")["id"], 27)
        self.assertEqual(self.client.get_product_by_symbol("ETH/USD")["id"], 27)
        self.assertEqual(self.client.get_product_by_symbol("BTCUSD.PERP")["id"], 1)
        self.assertEqual(self.client.get_product_by_symbol("DELTAIN:ETHUSD.P")["id"], 27)
        self.assertEqual(self.client.get_product_by_symbol("DELTA:ETHUSD.P")["id"], 27)

    def test_signature_generation(self):
        """Verify headers and signature concatenation formatting."""
        method = "POST"
        path = "/v2/orders"
        query_string = "product_ids=27"
        body = '{"product_id":27,"size":10}'
        
        with patch('time.time', return_value=1700000000):
            headers = self.client._generate_headers(method, path, query_string, body)
            
            # Expected signature calculation input
            expected_data = "POST1700000000/v2/orders?product_ids=27" + body
            expected_signature = hmac.new(
                self.api_secret.encode('utf-8'),
                expected_data.encode('utf-8'),
                hashlib.sha256
            ).hexdigest()
            
            self.assertEqual(headers["api-key"], self.api_key)
            self.assertEqual(headers["timestamp"], "1700000000")
            self.assertEqual(headers["signature"], expected_signature)

    def test_dynamic_sizing_math(self):
        """Verify math formula for contract sizing matches Delta's lot structure."""
        # Setup test parameters
        balance = 11.0 # USD
        leverage = 50
        buffer = 0.90 # 90%
        price = 1900.0 # ETH price
        contract_value = 0.01 # 1 lot = 0.01 ETH
        
        # Sizing logic from app.py
        buying_power = balance * leverage * buffer
        lot_value_usd = price * contract_value
        qty_lots = int(math.floor(buying_power / lot_value_usd))
        
        # Expected:
        # Buying power = 11.0 * 50 * 0.9 = 495 USD
        # Lot value = 1900 * 0.01 = 19 USD
        # Qty = floor(495 / 19) = floor(26.05) = 26 lots
        self.assertEqual(qty_lots, 26)
        
        # Test extreme case where balance is too small for 1 lot
        balance_small = 0.20
        buying_power_small = balance_small * leverage * buffer # 0.20 * 50 * 0.9 = 9.0 USD
        qty_lots_small = int(math.floor(buying_power_small / lot_value_usd)) # floor(9.0 / 19) = 0
        self.assertEqual(qty_lots_small, 0)

class TestWebhookEndpoints(unittest.TestCase):
    def setUp(self):
        import os
        os.environ["FLASK_ENV"] = "testing"
        os.environ["DATABASE_URL"] = "sqlite:///:memory:"
        os.environ["PASSPHRASE"] = "test_passphrase"
        os.environ["DEFAULT_LEVERAGE"] = "50"
        os.environ["BALANCE_BUFFER_PCT"] = "95"
        
        # Import app, config, and database models
        from app import app, db, Account, GlobalSetting
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
        
        self.app_context = app.app_context()
        self.app_context.push()
        
        db.session.rollback()
        db.session.close()
        db.session.expunge_all()
        db.create_all()
        
        # Add default passphrase
        db.session.add(GlobalSetting(key="passphrase", value="test_passphrase"))
        
        # Add a default active account
        self.default_account = Account(
            name="Test Account 1",
            api_key="key1",
            api_secret="secret1",
            leverage=50,
            balance_buffer_pct=95.0,
            is_active=True
        )
        db.session.add(self.default_account)
        db.session.commit()
        
        self.app = app.test_client()
        self.db = db
        self.Account = Account
        self.GlobalSetting = GlobalSetting
        
        # Mock the public symbol lookup to return consistent test values
        from app import public_delta_client
        public_delta_client.get_product_by_symbol = MagicMock(
            return_value={"symbol": "ETHUSD", "id": 27, "contract_value": "0.01"}
        )
        
    def tearDown(self):
        self.db.session.rollback()
        self.db.session.close()
        self.db.session.expunge_all()
        self.db.drop_all()
        self.app_context.pop()
        
    @patch('app.DeltaClient')
    def test_webhook_unauthorized(self, mock_client_class):
        # Setup basic mock for symbol lookup (uses public_delta_client)
        mock_client = mock_client_class.return_value
        mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27}

        # Test missing passphrase
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P"
        })
        self.assertEqual(response.status_code, 401)
        
        # Test invalid passphrase
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "wrong_passphrase"
        })
        self.assertEqual(response.status_code, 401)

    @patch('app.DeltaClient')
    def test_webhook_close_guards(self, mock_client_class):
        # Setup mocks
        mock_client = mock_client_class.return_value
        mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27}
        
        # Scenario 1: Receive close_long but position is SHORT (should ignore)
        mock_client.get_position.return_value = {"product_id": 27, "size": "-10", "side": "sell"}
        response = self.app.post('/webhook', json={
            "action": "close_long",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertEqual(len(results), 1)
        self.assertIn("Current position is not LONG, ignoring close_long", results[0]["message"])
        mock_client.place_order.assert_not_called()
        
        # Scenario 2: Receive close_short but position is LONG (should ignore)
        mock_client.get_position.return_value = {"product_id": 27, "size": "10", "side": "buy"}
        mock_client.place_order.reset_mock()
        response = self.app.post('/webhook', json={
            "action": "close_short",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertEqual(len(results), 1)
        self.assertIn("Current position is not SHORT, ignoring close_short", results[0]["message"])
        mock_client.place_order.assert_not_called()

        # Scenario 3: Receive close_long and position is LONG (should execute close)
        mock_client.get_position.return_value = {"product_id": 27, "size": "10", "side": "buy"}
        mock_client.place_order.reset_mock()
        mock_client.place_order.return_value = {"success": True, "result": {"id": 12345}}
        response = self.app.post('/webhook', json={
            "action": "close_long",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertTrue(results[0]["success"])
        mock_client.place_order.assert_called_once_with(
            product_id=27,
            size=10,
            side="sell",
            order_type="market_order",
            reduce_only=True
        )

    @patch('app.DeltaClient')
    def test_webhook_reversal_buy(self, mock_client_class):
        # Setup mocks
        mock_client = mock_client_class.return_value
        mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27, "contract_value": "0.01"}
        mock_client.get_position.return_value = {"product_id": 27, "size": "-10", "side": "sell"} # Currently SHORT
        mock_client.place_order.return_value = {"success": True, "result": {"id": 12345}}
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (11.0, "USD")
        
        # Send buy webhook (which triggers reversal of the short position)
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertTrue(results[0]["success"])
        
        # Verify first order was the close order (reduce_only=True)
        # Verify second order was the enter order
        calls = mock_client.place_order.call_args_list
        self.assertEqual(len(calls), 2)
        
        # Call 1: Close
        self.assertEqual(calls[0][1]["side"], "buy")
        self.assertEqual(calls[0][1]["size"], 10)
        self.assertTrue(calls[0][1]["reduce_only"])
        
        # Call 2: Enter
        # qty_lots = floor( (11.0 * 50 * 0.95) / (2000 * 0.01) ) = floor( 522.5 / 20 ) = 26 lots
        self.assertEqual(calls[1][1]["side"], "buy")
        self.assertEqual(calls[1][1]["size"], 26)
        self.assertFalse(calls[1][1]["reduce_only"])

    @patch('app.DeltaClient')
    def test_webhook_multi_account_replication(self, mock_client_class):
        # Add another active account, and one inactive account
        acc2 = self.Account(
            name="Test Account 2",
            api_key="key2",
            api_secret="secret2",
            leverage=20,
            balance_buffer_pct=50.0,
            is_active=True
        )
        acc_inactive = self.Account(
            name="Inactive Account",
            api_key="key3",
            api_secret="secret3",
            leverage=10,
            balance_buffer_pct=90.0,
            is_active=False
        )
        self.db.session.add(acc2)
        self.db.session.add(acc_inactive)
        self.db.session.commit()
        
        # Setup mock client instance
        mock_client = mock_client_class.return_value
        
        # Setup shared mocks
        mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27, "contract_value": "0.01"}
        mock_client.get_position.return_value = None # No open positions (no reversal needed)
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (10.0, "USD")
        mock_client.place_order.return_value = {"success": True, "result": {"id": 999}}
        
        # Trigger webhook
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertEqual(len(results), 2) # Should execute only on the 2 active accounts
        
        # Verify account names in results
        names = [res["name"] for res in results]
        self.assertIn("Test Account 1", names)
        self.assertIn("Test Account 2", names)
        self.assertNotIn("Inactive Account", names)
        
        # Verify place_order was called for each active account with its custom settings:
        # Account 1: leverage=50, buffer=95.0% -> Buying power = 10.0 * 50 * 0.95 = 475 USD
        # Lot value = 2000 * 0.01 = 20 USD. Qty = floor(475/20) = 23 lots
        # Account 2: leverage=20, buffer=50.0% -> Buying power = 10.0 * 20 * 0.50 = 100 USD
        # Lot value = 20 USD. Qty = floor(100/20) = 5 lots
        
        # Let's count how many times place_order was called on mock_client. It should be 2.
        calls = mock_client.place_order.call_args_list
        self.assertEqual(len(calls), 2)
        
        sizes = {call[1]["size"] for call in calls}
        self.assertIn(23, sizes)
        self.assertIn(5, sizes)

    @patch('app.DeltaClient')
    def test_webhook_fallback_when_db_empty(self, mock_client_class):
        # 1. Delete the default account from DB
        self.db.session.delete(self.default_account)
        self.db.session.commit()
        
        # 2. Mock environment variables on Config
        with patch('config.Config.API_KEY', 'env_fallback_key'), \
             patch('config.Config.API_SECRET', 'env_fallback_secret'):
            
            mock_client = mock_client_class.return_value
            mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27, "contract_value": "0.01"}
            mock_client.get_position.return_value = None
            mock_client.get_ticker.return_value = {"mark_price": "2000"}
            mock_client.get_available_balance.return_value = (10.0, "USD")
            mock_client.place_order.return_value = {"success": True, "result": {"id": 888}}
            
            response = self.app.post('/webhook', json={
                "action": "buy",
                "ticker": "ETHUSD.P",
                "passphrase": "test_passphrase"
            })
            
            self.assertEqual(response.status_code, 200)
            results = response.get_json()["results"]
            self.assertEqual(len(results), 1)
            self.assertEqual(results[0]["name"], "Environment Default")
            self.assertTrue(results[0]["success"])
            
            # Verify constructor of DeltaClient was called with the environment credentials
            mock_client_class.assert_called_with(
                api_key='env_fallback_key',
                api_secret='env_fallback_secret',
                base_url=Config.BASE_URL
            )
    def test_email_parsing_heuristics(self):
        """Verify email parsing correctly extracts ticker, action and quantity from different formats."""
        from app import parse_email_signal
        
        # Format 1: JSON body
        body_json = 'Some text before\n{\n  "action": "buy",\n  "ticker": "ETHUSD.P",\n  "quantity": 0.25\n}\nSome text after'
        subject = "Alert triggered"
        ticker, action, quantity = parse_email_signal(body_json, subject)
        self.assertEqual(ticker, "ETHUSD.P")
        self.assertEqual(action, "buy")
        self.assertEqual(quantity, 0.25)
        
        # Format 2: Key-value text
        body_kv = "Hello,\nTicker: BTCUSD.P\nAction: sell\nQty: 0.005\nThanks"
        ticker, action, quantity = parse_email_signal(body_kv, subject)
        self.assertEqual(ticker, "BTCUSD.P")
        self.assertEqual(action, "sell")
        self.assertEqual(quantity, 0.005)
        
        # Format 3: Subject line fallback
        body_empty = "This is a custom alert mail body."
        subject_buy = "Buy ETHUSD.P Alert"
        ticker, action, quantity = parse_email_signal(body_empty, subject_buy)
        self.assertEqual(ticker, "ETHUSD.P")
        self.assertEqual(action, "buy")
        self.assertIsNone(quantity)

    @patch('app.DeltaClient')
    def test_trade_logging_in_database(self, mock_client_class):
        """Verify that trade logs are correctly recorded in the SQLite database."""
        from app import TradeLog
        
        # Setup mock client
        mock_client = mock_client_class.return_value
        mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27, "contract_value": "0.01"}
        mock_client.get_position.return_value = None
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (10.0, "USD")
        mock_client.place_order.return_value = {"success": True, "result": {"id": 111}}
        
        # Count initial logs
        initial_count = TradeLog.query.count()
        
        # Trigger webhook
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(TradeLog.query.count(), initial_count + 1)
        
        latest_log = TradeLog.query.order_by(TradeLog.id.desc()).first()
        self.assertEqual(latest_log.ticker, "ETHUSD.P")
        self.assertEqual(latest_log.action, "buy")
        self.assertEqual(latest_log.status, "success")
        self.assertIn("Test Account 1: Success", latest_log.details)

    @patch('app.DeltaClient')
    def test_position_reconciliation_logic(self, mock_client_class):
        """Verify reconciliation checks skip matched positions and execute fallback for mismatched ones."""
        from app import check_position_matches_action
        
        mock_client = mock_client_class.return_value
        mock_client.get_product_by_symbol.return_value = {"symbol": "ETHUSD", "id": 27}
        
        account_data = {
            "api_key": "key1",
            "api_secret": "secret1"
        }
        
        # Scenario 1: Long position matches "buy" action
        mock_client.get_position.return_value = {"product_id": 27, "size": "10", "side": "buy"}
        self.assertTrue(check_position_matches_action(account_data, "ETHUSD.P", "buy"))
        
        # Scenario 2: Short position mismatches "buy" action
        mock_client.get_position.return_value = {"product_id": 27, "size": "-10", "side": "sell"}
        self.assertFalse(check_position_matches_action(account_data, "ETHUSD.P", "buy"))
        
        # Scenario 3: No position matches "close_long" action
        mock_client.get_position.return_value = None
        self.assertTrue(check_position_matches_action(account_data, "ETHUSD.P", "close_long"))

class TestPnLTracking(unittest.TestCase):
    def setUp(self):
        import os
        os.environ["FLASK_ENV"] = "testing"
        os.environ["DATABASE_URL"] = "sqlite:///:memory:"
        os.environ["PASSPHRASE"] = "test_passphrase"
        
        from app import app, db, Account
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
        
        self.app_context = app.app_context()
        self.app_context.push()
        
        db.session.rollback()
        db.session.close()
        db.session.expunge_all()
        db.create_all()
        
        # Add a test active account
        self.account = Account(
            name="PnL Test Account",
            api_key="pnl_key",
            api_secret="pnl_secret",
            leverage=10,
            balance_buffer_pct=50.0,
            is_active=True
        )
        db.session.add(self.account)
        db.session.commit()
        
        self.app = app.test_client()
        self.db = db

    def tearDown(self):
        self.db.session.rollback()
        self.db.session.close()
        self.db.session.expunge_all()
        self.db.drop_all()
        self.app_context.pop()

    @patch('delta_client.requests.get')
    def test_delta_client_get_available_balance(self, mock_get):
        """Test DeltaClient.get_available_balance with standard stablecoins and INR."""
        client = DeltaClient(api_key="key", api_secret="secret", base_url="https://api.delta.exchange")
        
        # Scenario A: Standard USDT/USDC wallet
        mock_response = MagicMock()
        mock_response.ok = True
        mock_response.json.return_value = {
            "success": True,
            "result": [
                {"asset_symbol": "USDT", "available_balance": "100.5"},
                {"asset_symbol": "USDC", "available_balance": "50.0"}
            ]
        }
        mock_get.return_value = mock_response
        balance, asset = client.get_available_balance()
        self.assertEqual(balance, 100.5)
        self.assertEqual(asset, "USDT")
        
        # Scenario B: INR wallet (Delta Exchange India) converted using 85.0 fixed rate
        mock_response_inr = MagicMock()
        mock_response_inr.ok = True
        mock_response_inr.json.return_value = {
            "success": True,
            "result": [
                {"asset_symbol": "INR", "available_balance": "8500.0"}
            ]
        }
        mock_get.return_value = mock_response_inr
        balance_inr, asset_inr = client.get_available_balance()
        self.assertAlmostEqual(balance_inr, 100.0, places=4)
        self.assertEqual(asset_inr, "USD")

    @patch('delta_client.requests.get')
    def test_delta_client_positions(self, mock_get):
        """Test DeltaClient methods get_open_positions and get_closed_positions."""
        client = DeltaClient(api_key="key", api_secret="secret", base_url="https://api.delta.exchange")
        
        # 1. Mock Open Positions response
        mock_open_response = MagicMock()
        mock_open_response.ok = True
        mock_open_response.json.return_value = {
            "success": True,
            "result": [
                {"product_id": 27, "size": "10", "side": "buy", "unrealized_pnl": "5.5"}
            ]
        }
        mock_get.return_value = mock_open_response
        
        open_pos = client.get_open_positions()
        self.assertEqual(len(open_pos), 1)
        self.assertEqual(open_pos[0]["product_id"], 27)
        self.assertEqual(open_pos[0]["unrealized_pnl"], "5.5")
        
        # 2. Mock Closed Positions (Orders History) response
        mock_closed_response = MagicMock()
        mock_closed_response.ok = True
        mock_closed_response.json.return_value = {
            "success": True,
            "result": [
                {
                    "product_id": 27,
                    "product": {"symbol": "ETHUSD", "contract_value": "0.01"},
                    "size": 10.0,
                    "side": "sell",
                    "average_fill_price": "1840.0",
                    "paid_commission": "0.05",
                    "state": "closed",
                    "created_at": "2026-06-11T18:10:07.453981Z",
                    "meta_data": {
                        "pnl": "-2.5",
                        "entry_price": "1850.0"
                    }
                }
            ]
        }
        mock_get.return_value = mock_closed_response
        
        closed_pos = client.get_closed_positions(limit=50)
        self.assertEqual(len(closed_pos), 1)
        self.assertEqual(closed_pos[0]["product_id"], 27)
        self.assertEqual(closed_pos[0]["realized_pnl"], -2.5)
        self.assertEqual(closed_pos[0]["side"], "LONG")
        self.assertEqual(closed_pos[0]["closed_at"], "2026-06-11T18:10:07.453981Z")

    @patch('app.DeltaClient')
    @patch('app.public_delta_client')
    def test_api_pnl_route(self, mock_public_client, mock_client_class):
        """Test that the /api/pnl endpoint aggregates and normalizes positions."""
        # Mock public product list
        mock_public_client.get_products.return_value = [
            {"id": 27, "symbol": "ETHUSD"}
        ]
        
        # Mock private client instance
        mock_private_client = mock_client_class.return_value
        
        # Mock open positions return
        mock_private_client.get_open_positions.return_value = [
            {
                "product_id": 27,
                "size": "10",
                "side": "buy",
                "entry_price": "1850",
                "mark_price": "1860",
                "unrealized_pnl": "10",
                "margin": "2",
                "leverage": "10"
            }
        ]
        
        # Mock closed positions return
        mock_private_client.get_closed_positions.return_value = [
            {
                "product_id": 27,
                "closed_size": "5",
                "side": "sell",
                "entry_price": "1850",
                "close_price": "1840",
                "realized_pnl": "-50",
                "closed_at": "1700000000"
            }
        ]
        
        response = self.app.get('/api/pnl')
        self.assertEqual(response.status_code, 200)
        
        data = response.get_json()
        self.assertIn("open", data)
        self.assertIn("closed", data)
        
        # Verify open position normalization
        open_list = data["open"]
        self.assertEqual(len(open_list), 1)
        self.assertEqual(open_list[0]["account_name"], "PnL Test Account")
        self.assertEqual(open_list[0]["symbol"], "ETHUSD")
        self.assertEqual(open_list[0]["side"], "LONG")
        self.assertEqual(open_list[0]["size"], 10.0)
        self.assertEqual(open_list[0]["unrealized_pnl"], 10.0)
        
        # Verify closed trade history normalization
        closed_list = data["closed"]
        self.assertEqual(len(closed_list), 1)
        self.assertEqual(closed_list[0]["account_name"], "PnL Test Account")
        self.assertEqual(closed_list[0]["symbol"], "ETHUSD")
        self.assertEqual(closed_list[0]["side"], "SHORT")
        self.assertEqual(closed_list[0]["closed_size"], 5.0)
        self.assertEqual(closed_list[0]["realized_pnl"], -50.0)
        self.assertIn("fees", closed_list[0])
        self.assertIn("net_pnl", closed_list[0])
        self.assertAlmostEqual(closed_list[0]["fees"], 0.09225, places=5)
        self.assertAlmostEqual(closed_list[0]["net_pnl"], -50.09225, places=5)

    @patch('app.DeltaClient')
    @patch('app.public_delta_client')
    def test_export_journal_csv(self, mock_public_client, mock_client_class):
        """Test that the /api/journal/export endpoint generates a valid CSV file download."""
        # Mock public product list
        mock_public_client.get_products.return_value = [
            {"id": 27, "symbol": "ETHUSD", "contract_value": "0.01"}
        ]
        
        # Mock private client instance
        mock_private_client = mock_client_class.return_value
        
        # Mock closed positions return
        mock_private_client.get_closed_positions.return_value = [
            {
                "product_id": 27,
                "closed_size": "10",
                "side": "buy",
                "entry_price": "2000",
                "close_price": "2020",
                "realized_pnl": "2.0",
                "closed_at": "1700000000"
            }
        ]
        
        response = self.app.get('/api/journal/export')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.content_type, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("attachment; filename=trading_journal.xlsx", response.headers["Content-Disposition"])
        
        # Load Excel output
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        ws = wb.active
        self.assertEqual(ws.title, "Trading Journal")
        
        # Verify headers exist
        self.assertEqual(ws.cell(row=1, column=1).value, "CLOSED TIME (UTC)")
        self.assertEqual(ws.cell(row=1, column=2).value, "ACCOUNT NAME")
        self.assertEqual(ws.cell(row=1, column=3).value, "SYMBOL")
        self.assertEqual(ws.cell(row=1, column=8).value, "GROSS PNL (USD)")
        self.assertEqual(ws.cell(row=1, column=9).value, "FEES & COMMISSION (USD)")
        self.assertEqual(ws.cell(row=1, column=10).value, "NET PNL (USD)")
        
        # Verify data row values
        self.assertEqual(ws.cell(row=2, column=2).value, "PnL Test Account")
        self.assertEqual(ws.cell(row=2, column=3).value, "ETHUSD")
        self.assertEqual(ws.cell(row=2, column=4).value, "LONG")
        self.assertEqual(ws.cell(row=2, column=5).value, 10.0)
        self.assertEqual(ws.cell(row=2, column=8).value, 2.0)
        
        # Expected Fees: entry_notional = 2000 * 10 * 0.01 = 200. exit_notional = 2020 * 10 * 0.01 = 202.
        # total_notional = 402. fees = 402 * 0.0005 = 0.2010.
        # Expected Net PnL: 2.0 - 0.201 = 1.7990.
        self.assertAlmostEqual(ws.cell(row=2, column=9).value, 0.2010)
        self.assertAlmostEqual(ws.cell(row=2, column=10).value, 1.7990)
        
        # Verify Total row
        self.assertEqual(ws.cell(row=3, column=1).value, "TOTAL")
        self.assertEqual(ws.cell(row=3, column=8).value, "=SUM(H2:H2)")
        self.assertEqual(ws.cell(row=3, column=9).value, "=SUM(I2:I2)")
        self.assertEqual(ws.cell(row=3, column=10).value, "=SUM(J2:J2)")

class TestSizingModels(unittest.TestCase):
    def setUp(self):
        import os
        os.environ["FLASK_ENV"] = "testing"
        os.environ["DATABASE_URL"] = "sqlite:///:memory:"
        os.environ["PASSPHRASE"] = "test_passphrase"
        
        from app import app, db, Account, GlobalSetting
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
        
        self.app_context = app.app_context()
        self.app_context.push()
        
        db.session.rollback()
        db.session.close()
        db.session.expunge_all()
        db.create_all()
        
        db.session.add(GlobalSetting(key="passphrase", value="test_passphrase"))
        db.session.commit()
        
        self.app = app.test_client()
        self.db = db
        self.Account = Account

        # Mock public symbol lookup
        from app import public_delta_client
        public_delta_client.get_product_by_symbol = MagicMock(
            return_value={"symbol": "ETHUSD", "id": 27, "contract_value": "0.01"}
        )

    def tearDown(self):
        self.db.session.rollback()
        self.db.session.close()
        self.db.session.expunge_all()
        self.db.drop_all()
        self.app_context.pop()

    @patch('app.DeltaClient')
    def test_fixed_margin_sizing_success(self, mock_client_class):
        """Verify that sizing_type='fixed' computes quantity using fixed margin rather than percentage."""
        # 1. Create account with fixed sizing
        acc = self.Account(
            name="Fixed Account Success",
            api_key="key_fixed",
            api_secret="secret_fixed",
            leverage=50,
            sizing_type="fixed",
            fixed_amount=8.0, # Commit exactly $8.00 of margin
            is_active=True
        )
        self.db.session.add(acc)
        self.db.session.commit()

        # 2. Setup mock client response
        mock_client = mock_client_class.return_value
        mock_client.get_position.return_value = None
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (10.0, "USD") # Available balance is $10.00
        mock_client.place_order.return_value = {"success": True, "result": {"id": 111}}
        
        # 3. Post webhook
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertTrue(results[0]["success"])
        
        # Sizing math:
        # Buying power = fixed_amount * leverage = 8.0 * 50 = 400 USD
        # Lot value = 2000 * 0.01 = 20 USD
        # Qty = floor(400 / 20) = 20 lots
        mock_client.place_order.assert_called_once_with(
            product_id=27,
            size=20,
            side="buy",
            order_type="market_order",
            reduce_only=False
        )

    @patch('app.DeltaClient')
    def test_fixed_margin_sizing_insufficient_balance(self, mock_client_class):
        """Verify that if fixed margin amount exceeds available balance, the trade is aborted."""
        # 1. Create account with fixed sizing where fixed_amount is greater than balance
        acc = self.Account(
            name="Fixed Account Insufficient",
            api_key="key_fixed2",
            api_secret="secret_fixed2",
            leverage=50,
            sizing_type="fixed",
            fixed_amount=15.0, # Requires $15.00 margin
            is_active=True
        )
        self.db.session.add(acc)
        self.db.session.commit()

        # 2. Setup mock client response
        mock_client = mock_client_class.return_value
        mock_client.get_position.return_value = None
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (10.0, "USD") # Only has $10.00
        
        # 3. Post webhook
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase"
        })
        
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertFalse(results[0]["success"])
        self.assertIn("Insufficient balance (need 15.0 USD, have 10.0 USD)", results[0]["message"])
        
        # Assert order was NOT placed
        mock_client.place_order.assert_not_called()

    @patch('app.DeltaClient')
    def test_payload_quantity_sizing_success(self, mock_client_class):
        """Verify that passing quantity in webhook payload overrides default sizing and places order correctly."""
        # 1. Create active account
        acc = self.Account(
            name="Payload Qty Account",
            api_key="key_payload",
            api_secret="secret_payload",
            leverage=50,
            is_active=True
        )
        self.db.session.add(acc)
        self.db.session.commit()

        # 2. Setup mock client response
        mock_client = mock_client_class.return_value
        mock_client.get_position.return_value = None
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (10.0, "USD") # Available balance is $10.00
        mock_client.place_order.return_value = {"success": True, "result": {"id": 222}}
        
        # 3. Post webhook with quantity parameter
        # ETHUSD.P contract value is 0.01. So quantity of 0.25 ETH = 25 contracts/lots
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase",
            "quantity": 0.25
        })
        
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertTrue(results[0]["success"])
        
        # Verify place_order was called with 25 lots (0.25 / 0.01)
        mock_client.place_order.assert_called_once_with(
            product_id=27,
            size=25,
            side="buy",
            order_type="market_order",
            reduce_only=False
        )

    @patch('app.DeltaClient')
    def test_payload_quantity_insufficient_margin(self, mock_client_class):
        """Verify that if quantity from payload requires more margin than available, the trade is aborted."""
        # 1. Create active account
        acc = self.Account(
            name="Payload Qty Overlimit",
            api_key="key_payload2",
            api_secret="secret_payload2",
            leverage=10, # low leverage to increase margin required
            is_active=True
        )
        self.db.session.add(acc)
        self.db.session.commit()

        # 2. Setup mock client response
        mock_client = mock_client_class.return_value
        mock_client.get_position.return_value = None
        mock_client.get_ticker.return_value = {"mark_price": "2000"}
        mock_client.get_available_balance.return_value = (10.0, "USD") # Available balance is $10.00
        
        # 3. Post webhook requesting huge quantity
        # 0.1 ETH @ $2000 = $200 position value.
        # Leverage = 10x -> Required margin = $20.00, which exceeds balance of $10.00
        response = self.app.post('/webhook', json={
            "action": "buy",
            "ticker": "ETHUSD.P",
            "passphrase": "test_passphrase",
            "quantity": 0.10
        })
        
        self.assertEqual(response.status_code, 200)
        results = response.get_json()["results"]
        self.assertFalse(results[0]["success"])
        self.assertIn("Insufficient balance", results[0]["message"])
        
        # Assert order was NOT placed
        mock_client.place_order.assert_not_called()

class TestNotificationSettings(unittest.TestCase):
    def setUp(self):
        import os
        os.environ["FLASK_ENV"] = "testing"
        from app import app, db, GlobalSetting
        self.GlobalSetting = GlobalSetting
        self.db = db
        
        app.config['TESTING'] = True
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
        self.app = app.test_client()
        
        # Bind the app context
        self.app_context = app.app_context()
        self.app_context.push()
        
        # Clean up any session leftovers
        db.session.rollback()
        db.session.close()
        db.session.expunge_all()
        db.create_all()
        
        # Clear existing passphrase to ensure clean test environment
        GlobalSetting.query.filter_by(key="passphrase").delete()
        db.session.add(GlobalSetting(key="passphrase", value="test_passphrase"))
        db.session.commit()

    def tearDown(self):
        self.db.session.remove()
        self.db.drop_all()
        self.app_context.pop()

    def test_get_settings(self):
        # Add notification settings
        self.db.session.add(self.GlobalSetting(key="telegram_enabled", value="true"))
        self.db.session.add(self.GlobalSetting(key="telegram_token", value="my_token"))
        self.db.session.add(self.GlobalSetting(key="telegram_chat_id", value="123456"))
        self.db.session.add(self.GlobalSetting(key="discord_enabled", value="true"))
        self.db.session.add(self.GlobalSetting(key="discord_webhook_url", value="my_discord"))
        self.db.session.commit()

        response = self.app.get('/api/settings')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertEqual(data["passphrase"], "test_passphrase")
        self.assertEqual(data["telegram_enabled"], "true")
        self.assertEqual(data["telegram_token"], "my_token")
        self.assertEqual(data["telegram_chat_id"], "123456")
        self.assertEqual(data["discord_enabled"], "true")
        self.assertEqual(data["discord_webhook_url"], "my_discord")

    def test_save_settings(self):
        payload = {
            "passphrase": "new_passphrase",
            "telegram_enabled": "true",
            "telegram_token": "my_new_token",
            "telegram_chat_id": "987654",
            "discord_enabled": "false",
            "discord_webhook_url": "my_new_discord"
        }
        response = self.app.post('/api/settings', json=payload)
        self.assertEqual(response.status_code, 200)
        
        # Assert database was updated
        passphrase = self.GlobalSetting.query.filter_by(key="passphrase").first()
        self.assertEqual(passphrase.value, "new_passphrase")
        
        tg_token = self.GlobalSetting.query.filter_by(key="telegram_token").first()
        self.assertEqual(tg_token.value, "my_new_token")

    def test_notifications_test_endpoint(self):
        # Testing test endpoint returns success response
        response = self.app.post('/api/notifications/test')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["status"], "success")

class TestWebhookPlayground(unittest.TestCase):
    def setUp(self):
        import os
        os.environ["FLASK_ENV"] = "testing"
        from app import app, db, Account, GlobalSetting
        self.Account = Account
        self.GlobalSetting = GlobalSetting
        self.db = db
        
        app.config['TESTING'] = True
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
        self.app = app.test_client()
        
        # Bind the app context
        self.app_context = app.app_context()
        self.app_context.push()
        
        # Initialize tables
        db.session.rollback()
        db.session.close()
        db.session.expunge_all()
        db.create_all()
        
        # Seed test passphrase
        db.session.add(GlobalSetting(key="passphrase", value="test_passphrase"))
        
        # Seed an active account
        self.acc = Account(
            name="Playground Test Acc",
            api_key="play_key",
            api_secret="play_secret",
            leverage=20,
            balance_buffer_pct=50.0,
            sizing_type="percentage",
            is_active=True
        )
        db.session.add(self.acc)
        db.session.commit()

    def tearDown(self):
        self.db.session.remove()
        self.db.drop_all()
        self.app_context.pop()

    @patch('app.public_delta_client')
    def test_simulate_webhook_dry_run_success(self, mock_public_client):
        # Mock product resolution
        mock_public_client.get_product_by_symbol.return_value = {
            "id": 27,
            "symbol": "ETHUSD",
            "contract_value": "0.01"
        }
        
        payload = {
            "ticker": "ETHUSD.P",
            "action": "buy",
            "passphrase": "test_passphrase"
        }
        response = self.app.post('/api/simulate-webhook', json=payload)
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        self.assertTrue(data["simulation"])
        self.assertEqual(data["ticker"], "ETHUSD.P")
        self.assertEqual(data["results"][0]["account"], "Playground Test Acc")
        self.assertTrue(data["results"][0]["success"])
        self.assertIn("Simulated successfully", data["results"][0]["message"])

    @patch('app.public_delta_client')
    def test_simulate_webhook_invalid_passphrase(self, mock_public_client):
        payload = {
            "ticker": "ETHUSD.P",
            "action": "buy",
            "passphrase": "wrong_passphrase"
        }
        response = self.app.post('/api/simulate-webhook', json=payload)
        self.assertEqual(response.status_code, 401)
        self.assertEqual(response.get_json()["message"], "Invalid passphrase")

class TestCircuitBreaker(unittest.TestCase):
    def setUp(self):
        import os
        os.environ["FLASK_ENV"] = "testing"
        from app import app, db, Account, GlobalSetting
        self.Account = Account
        self.GlobalSetting = GlobalSetting
        self.db = db
        
        app.config['TESTING'] = True
        app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///:memory:'
        self.app = app.test_client()
        
        # Bind the app context
        self.app_context = app.app_context()
        self.app_context.push()
        
        # Initialize tables
        db.session.rollback()
        db.session.close()
        db.session.expunge_all()
        db.create_all()
        
        # Seed test passphrase
        db.session.add(GlobalSetting(key="passphrase", value="test_passphrase"))
        
        # Seed an active account with a daily loss limit
        self.acc = Account(
            name="Breaker Test Acc",
            api_key="key1",
            api_secret="secret1",
            leverage=20,
            balance_buffer_pct=50.0,
            sizing_type="percentage",
            daily_loss_limit=10.0,
            is_circuit_broken=False,
            is_active=True
        )
        db.session.add(self.acc)
        db.session.commit()

    def tearDown(self):
        self.db.session.remove()
        self.db.drop_all()
        self.app_context.pop()

    @patch('app.DeltaClient')
    @patch('app.public_delta_client')
    def test_circuit_breaker_tripping(self, mock_public_client, mock_client_class):
        # Mock product resolution
        mock_public_client.get_product_by_symbol.return_value = {
            "id": 27,
            "symbol": "ETHUSD",
            "contract_value": "0.01"
        }
        
        # Mock private client instance
        mock_private_client = mock_client_class.return_value
        
        # Mock get_available_balance
        mock_private_client.get_available_balance.return_value = (100.0, "USD")
        
        # Mock get_ticker
        mock_private_client.get_ticker.return_value = {"mark_price": "2000.0", "last_price": "2000.0"}
        
        # Mock get_open_positions
        mock_private_client.get_open_positions.return_value = [
            {"product_id": 27, "size": "10", "side": "buy"}
        ]
        
        # Mock get_closed_positions to return trades closed today with a net loss of 15.0 USD
        import datetime
        now_utc = datetime.datetime.now(datetime.timezone.utc)
        mock_private_client.get_closed_positions.return_value = [
            {
                "product_id": 27,
                "closed_size": "5",
                "side": "sell",
                "entry_price": "1850",
                "close_price": "1840",
                "realized_pnl": "-15.0",
                "fee": "0.0",
                "closed_at": now_utc.isoformat()
            }
        ]
        
        # Send a webhook to trigger execute_trades_background
        payload = {
            "ticker": "ETHUSD.P",
            "action": "buy",
            "passphrase": "test_passphrase"
        }
        response = self.app.post('/webhook', json=payload)
        self.assertEqual(response.status_code, 200)
        
        # Verify account is now circuit broken in database
        self.db.session.expire_all()
        updated_acc = self.Account.query.get(self.acc.id)
        self.assertTrue(updated_acc.is_circuit_broken)
        
        # Assert place_order was called with reduce_only=True to close the open position
        mock_private_client.place_order.assert_any_call(
            product_id=27,
            size=10,
            side="sell",
            order_type="market_order",
            reduce_only=True
        )

    @patch('app.DeltaClient')
    @patch('app.public_delta_client')
    def test_already_circuit_broken_halts_trades(self, mock_public_client, mock_client_class):
        # Set is_circuit_broken to True
        self.acc.is_circuit_broken = True
        self.db.session.commit()
        
        # Mock product resolution
        mock_public_client.get_product_by_symbol.return_value = {
            "id": 27,
            "symbol": "ETHUSD",
            "contract_value": "0.01"
        }
        
        # Mock private client instance
        mock_private_client = mock_client_class.return_value
        
        payload = {
            "ticker": "ETHUSD.P",
            "action": "buy",
            "passphrase": "test_passphrase"
        }
        response = self.app.post('/webhook', json=payload)
        self.assertEqual(response.status_code, 200)
        
        # Verify order was not placed since trading is halted
        mock_private_client.place_order.assert_not_called()

    def test_reset_breaker_endpoint(self):
        # Make the account broken
        self.acc.is_circuit_broken = True
        self.db.session.commit()
        
        response = self.app.post(f'/api/accounts/{self.acc.id}/reset-breaker')
        self.assertEqual(response.status_code, 200)
        
        # Verify account is no longer circuit broken in database
        updated_acc = self.Account.query.get(self.acc.id)
        self.assertFalse(updated_acc.is_circuit_broken)

if __name__ == "__main__":
    unittest.main()
